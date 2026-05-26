"""Export a LessonPackage to a structured Excel workbook (file or bytes)."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schemas import LessonPackage

HEADER_BG = "4472C4"
HEADER_FG = "FFFFFF"
ALT_BG    = "DCE6F1"


def _header(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _wrap(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_workbook(package: LessonPackage) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Session Script ─────────────────────────────────────────────
    ws = wb.create_sheet("Session Script")
    cols = ["Section", "Round Label", "Speaker", "Line", "Is Action"]
    _header(ws, 1, cols)
    ss = package.session_script
    if ss:
        # Intro lines
        for line in ss.intro:
            ws.append(["Intro", "—", line.speaker, line.text, str(line.is_action)])
        # Round lines
        for rnd in ss.rounds:
            for line in rnd.dialog:
                ws.append([
                    f"Round {rnd.round_number}",
                    rnd.round_label,
                    line.speaker,
                    line.text,
                    str(line.is_action),
                ])
                # Color action rows
                if line.is_action:
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    _autowidth(ws)
    _wrap(ws)

    # ── Sheet 2: Practice Scenarios ────────────────────────────────────────
    ws = wb.create_sheet("Practice Scenarios")
    cols = ["ID", "Lesson ID", "Title", "Setup", "Nessa Prompt",
            "Target Response", "Small Hint", "Big Hint",
            "Example Correct Response", "Difficulty", "Tags"]
    _header(ws, 1, cols)
    for i, sc in enumerate(package.scenarios):
        fill = PatternFill("solid", fgColor=ALT_BG) if i % 2 else None
        ws.append([
            sc.scenario_id,
            sc.lesson_id,
            sc.title,
            sc.setup,
            sc.nessa_prompt,
            sc.target_response,
            sc.small_hint,
            sc.big_hint,
            sc.example_correct_response,
            sc.difficulty,
            ", ".join(sc.tags),
        ])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    _autowidth(ws)
    _wrap(ws)

    # ── Sheet 3: Scenario Image ─────────────────────────────────────────────
    ws = wb.create_sheet("Scenario Image")
    cols = ["Scenario ID", "Image Prompt", "Style", "Key Elements",
            "Child Description", "Color Palette"]
    _header(ws, 1, cols)
    for i, img in enumerate(package.scenario_images):
        fill = PatternFill("solid", fgColor=ALT_BG) if i % 2 else None
        ws.append([
            img.scenario_id,
            img.image_prompt,
            img.style,
            "\n".join(img.key_elements),
            img.child_description,
            img.color_palette or "",
        ])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    _autowidth(ws)
    _wrap(ws)

    # ── Sheet 4: QA Review ──────────────────────────────────────────────────
    ws = wb.create_sheet("QA Review")
    cols = ["Category", "Check Item", "Status", "Notes"]
    _header(ws, 1, cols)
    if package.qa_review:
        qa = package.qa_review
        all_items = (
            qa.prompt_quality + qa.scenario_clarity + qa.image_scenario_match
            + qa.description_length + qa.tool_use_rules
        )
        for item in all_items:
            color = {"pass": "C6EFCE", "fail": "FFC7CE", "warning": "FFEB9C"}.get(
                item.status.lower(), "FFFFFF"
            )
            ws.append([item.category, item.item, item.status, item.notes])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)

        summary_row = ws.max_row + 2
        ws.cell(summary_row, 1, "OVERALL").font = Font(bold=True)
        overall_color = "C6EFCE" if qa.overall_pass else "FFC7CE"
        ws.cell(summary_row, 3, "PASS" if qa.overall_pass else "FAIL").fill = PatternFill(
            "solid", fgColor=overall_color
        )
        ws.cell(summary_row, 4, qa.reviewer_notes)
    _autowidth(ws)
    _wrap(ws)

    return wb


def export(package: LessonPackage, output_path: Optional[str] = None) -> Path:
    wb = _build_workbook(package)
    lesson_id = package.session_script.lesson_id if package.session_script else "lesson"
    out = Path(output_path or f"{lesson_id}_package.xlsx")
    wb.save(out)
    return out


def export_bytes(package: LessonPackage) -> bytes:
    wb = _build_workbook(package)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
